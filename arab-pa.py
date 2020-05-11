from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from time import sleep
from bs

number = 1
url = 'http://www.arab-pa.org/en/shared/Publishers.aspx'
driver = webdriver.Chrome()
driver.get(url)
row = 1
while number <= 135:
    wb = load_workbook("arab-pa.xlsx")
    sheet = wb.active
    containers = driver.find_elements_by_css_selector(".pricing-desc")
    for container in containers:
        try:
            publisher = container.find_element_by_css_selector("div.pricing-title h3").text
        except:
            punlisher = ""
        try:
            name = container.find_element_by_css_selector(".iconlist-color li:nth-child(1)").text
        except:
            name = ""
        try:
            job = container.find_element_by_css_selector(".iconlist-color li:nth-child(2)").text
        except:
            job = ""
        try:
            country = container.find_element_by_css_selector(".iconlist-color li:nth-child(3)").text
        except:
            country = ""
        try:
            phone = container.find_element_by_css_selector(".iconlist-color li:nth-child(4)").text
        except:
            phone = ""
        try:
            fax = container.find_element_by_css_selector(".iconlist-color li:nth-child(5)").text
        except:
            fax = ""
        try:
            mobile = container.find_element_by_css_selector(".iconlist-color li:nth-child(6)").text
        except:
            mobile = ""
        try:
            email = container.find_element_by_css_selector(".iconlist-color li:nth-child(7)").text
        except:
            email = ""
        try:
            website = container.find_element_by_css_selector(".iconlist-color li:nth-child(8)").text
        except:
            website = ""
        sheet.cell(row=row, column=1, value=publisher)
        sheet.cell(row=row, column=2, value=name)
        sheet.cell(row=row, column=3, value=job)
        sheet.cell(row=row, column=4, value=country)
        sheet.cell(row=row, column=5, value=phone)
        sheet.cell(row=row, column=6, value=fax)
        sheet.cell(row=row, column=7, value=mobile)
        sheet.cell(row=row, column=8, value=email)
        sheet.cell(row=row, column=9, value=website)
        row+=1
    wb.save("arab-pa.xlsx")
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//table[@id="ctl00_ContentPlaceHolder1_GridView1"]//table/tbody/tr//td/span//following::td/a'))).click()
    print(number)
    number += 1
driver.quit()